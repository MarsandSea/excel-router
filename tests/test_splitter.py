"""core/splitter.py 的集成测试：表头识别、列枚举、跨文件合并、二级拆分、格式保留。"""
import os

import openpyxl
from openpyxl.styles import Font, PatternFill
import pytest

from core.splitter import (
    detect_header_row_auto, find_header_row, resolve_header_row,
    list_columns, list_values, run_split,
    _dedupe_path, _safe_sheet_title,
)


# ---------- 样本构造 ----------

HEADERS = ["工号", "姓名", "部门", "城市", "金额"]


def _make_book(path, rows, title="月度报表"):
    """表头在第2行、第1行是大标题（用于测自动表头识别）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = title
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
    for i, r in enumerate(rows, start=3):
        for c, v in enumerate(r, 1):
            ws.cell(row=i, column=c, value=v)
    wb.save(path)


def _count_rows(path, start=3):
    wb = openpyxl.load_workbook(path)
    total = 0
    for ws in wb.worksheets:
        for r in range(start, ws.max_row + 1):
            if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, ws.max_column + 1)):
                total += 1
    return total


@pytest.fixture
def two_books(tmp_path):
    inp = tmp_path / "in"
    out = tmp_path / "out"
    inp.mkdir()
    out.mkdir()
    _make_book(inp / "A.xlsx", [
        ["001", "张三", "销售部", "北京", 100],
        ["002", "李四", "销售部", "北京", 200],
        ["003", "王五", "销售部", "上海", 150],
        ["004", "赵六", "技术部", "北京", 300],
        ["", "合计", "", "", 750],
    ])
    _make_book(inp / "B.xlsx", [
        ["005", "钱七", "销售部", "广州", 120],
        ["006", "孙八", "技术部", "深圳", 220],
    ])
    return str(inp), str(out)


def _base_cfg(inp, out, **over):
    cfg = {
        "input_path": inp, "output_path": out,
        "header_mode": "auto", "header_row": 1,
        "grid_keys": [], "id_keys": [],
        "split_column": "部门", "selected_values": [],
        "person_column": "", "to_person": False, "person_file_filter": [],
        "value_alias_map": {},
        "skip_values": ["合计", "小计", "总计", ""],
        "merge_across_files": True, "make_zip": True, "exact_match": True,
        "preserve_format": True, "auto_open_output": False,
    }
    cfg.update(over)
    return cfg


# ---------- 表头识别 ----------

def test_detect_header_row_auto():
    rows = [["月度报表", None, None], ["工号", "姓名", "部门"], ["001", "张三", "销售部"]]
    assert detect_header_row_auto(rows) == 2


def test_resolve_header_modes():
    rows = [["标题"], ["网格", "工号", "姓名"], ["A", "001", "张三"]]
    assert resolve_header_row(rows, {"header_mode": "auto"}) == 2
    assert resolve_header_row(rows, {"header_mode": "row", "header_row": 2}) == 2
    assert resolve_header_row(
        rows, {"header_mode": "keyword", "grid_keys": ["网格"], "id_keys": ["工号"]}) == 2
    assert find_header_row(rows, ["不存在"], ["工号"]) == -1


# ---------- 列名 / 取值枚举 ----------

def test_list_columns_and_values(two_books):
    inp, _ = two_books
    cfg = {"header_mode": "auto"}
    a = os.path.join(inp, "A.xlsx")
    assert list_columns(a, cfg) == HEADERS
    assert list_values(a, cfg, "部门") == ["技术部", "销售部"]   # 合计行被跳过


# ---------- 跨文件合并（核心）----------

def test_merge_across_files(two_books):
    inp, out = two_books
    res = run_split(_base_cfg(inp, out), log_fn=lambda m: None)
    sales = os.path.join(res, "销售部.xlsx")
    tech = os.path.join(res, "技术部.xlsx")
    assert os.path.exists(sales) and os.path.exists(tech)
    # A 的 3 行销售部 + B 的 1 行销售部 = 4（证明合并而非覆盖）
    assert _count_rows(sales) == 4
    assert _count_rows(tech) == 2


def test_no_merge_keeps_separate(two_books):
    inp, out = two_books
    res = run_split(_base_cfg(inp, out, merge_across_files=False), log_fn=lambda m: None)
    # 不合并：销售部/汇总 下按源文件分文件
    folder = os.path.join(res, "销售部", "汇总")
    assert os.path.isdir(folder)
    files = [f for f in os.listdir(folder) if f.endswith(".xlsx")]
    assert len(files) == 2   # A、B 各一个
    assert os.path.exists(os.path.join(res, "销售部.zip"))   # 每主取值打包


def test_single_file_input(two_books):
    inp, _ = two_books
    # 输入单个文件 -> 扁平输出，无文件夹、无 ZIP
    one = os.path.join(inp, "A.xlsx")
    out = os.path.join(os.path.dirname(inp), "single_out")
    os.makedirs(out, exist_ok=True)
    res = run_split(_base_cfg(one, out), log_fn=lambda m: None)
    assert _count_rows(os.path.join(res, "销售部.xlsx")) == 3   # A 表销售部 3 行
    assert _count_rows(os.path.join(res, "技术部.xlsx")) == 1
    assert not os.path.isdir(os.path.join(res, "销售部"))
    assert not os.path.exists(os.path.join(res, "销售部.zip"))


# ---------- 格式保留 ----------

def test_header_format_preserved(two_books):
    inp, out = two_books
    res = run_split(_base_cfg(inp, out), log_fn=lambda m: None)
    wb = openpyxl.load_workbook(os.path.join(res, "销售部.xlsx"))
    ws = wb.active
    assert ws["A2"].value == "工号"
    assert ws["A2"].font.bold is True


# ---------- 到人（双产出 + 范围过滤）----------

def test_to_person_all(two_books):
    inp, out = two_books
    res = run_split(_base_cfg(inp, out, to_person=True, person_column="姓名",
                              merge_across_files=False), log_fn=lambda m: None)
    # 汇总照常产出
    assert os.path.isdir(os.path.join(res, "销售部", "汇总"))
    # 到人：销售部里 A 的张三/李四/王五 + B 的钱七 都各一个
    for name in ("张三", "李四", "王五", "钱七"):
        assert os.path.exists(os.path.join(res, "销售部", "到人", f"{name}.xlsx")), name
    # 技术部到人：A 赵六 + B 孙八
    assert os.path.exists(os.path.join(res, "技术部", "到人", "赵六.xlsx"))
    assert os.path.exists(os.path.join(res, "技术部", "到人", "孙八.xlsx"))


def test_to_person_scope_filter(two_books):
    inp, out = two_books
    # 仅文件名含 "A" 的表参与到人
    res = run_split(_base_cfg(inp, out, to_person=True, person_column="姓名",
                              person_file_filter=["A"], merge_across_files=False),
                    log_fn=lambda m: None)
    assert os.path.exists(os.path.join(res, "销售部", "到人", "张三.xlsx"))   # A 表
    assert not os.path.exists(os.path.join(res, "销售部", "到人", "钱七.xlsx"))  # B 表，不在范围
    # 但 B 表仍进了汇总
    assert os.path.exists(os.path.join(res, "销售部", "汇总", "B.xlsx"))


# ---------- 只拆部分取值 ----------

def test_selected_values_subset(two_books):
    inp, out = two_books
    res = run_split(_base_cfg(inp, out, selected_values=["销售部"]), log_fn=lambda m: None)
    assert os.path.exists(os.path.join(res, "销售部.xlsx"))
    assert not os.path.exists(os.path.join(res, "技术部.xlsx"))


# ---------- 取值归并映射 ----------

def test_value_alias_map(tmp_path):
    inp = tmp_path / "in"; out = tmp_path / "out"
    inp.mkdir(); out.mkdir()
    _make_book(inp / "C.xlsx", [
        ["001", "张三", "销售部", "北京", 100],
        ["002", "李四", "销售一部", "上海", 200],
    ])
    cfg = _base_cfg(str(inp), str(out),
                    value_alias_map={"销售": ["销售部", "销售一部"]})
    res = run_split(cfg, log_fn=lambda m: None)
    merged = os.path.join(res, "销售.xlsx")
    assert os.path.exists(merged)
    assert _count_rows(merged) == 2                   # 两个别名归并成一组


# ---------- 对抗式审查回归用例 ----------

def _make_book_custom(path, headers, rows, title="报表", sheet="Sheet1"):
    """可指定列头/行/sheet 名的样本（表头在第2行）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws["A1"] = title
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    for i, r in enumerate(rows, start=3):
        for c, v in enumerate(r, 1):
            ws.cell(row=i, column=c, value=v)
    wb.save(path)


def test_filename_collision_no_overwrite(tmp_path):
    """F1：两个净化后同名（销售_部）但不同的取值不能互相覆盖。"""
    import glob
    inp = tmp_path / "in"; out = tmp_path / "out"
    inp.mkdir(); out.mkdir()
    _make_book_custom(inp / "X.xlsx", HEADERS, [
        ["001", "张三", "销售/部", "北京", 100],
        ["002", "李四", "销售/部", "北京", 200],
        ["003", "王五", "销售:部", "上海", 150],   # 净化后同为「销售_部」
    ])
    res = run_split(_base_cfg(str(inp / "X.xlsx"), str(out)), log_fn=lambda m: None)
    files = glob.glob(os.path.join(res, "销售_部*.xlsx"))
    assert len(files) == 2, files                       # 两份并存（销售_部.xlsx + 销售_部(2).xlsx）
    assert sum(_count_rows(f) for f in files) == 3       # 3 行数据一行不丢


def test_skip_leftover_temp_file(two_books):
    """F2：旧版可能残留在输入目录的 *__tmp__.xlsx 不应被当成正式输入处理。"""
    inp, out = two_books
    _make_book_custom(os.path.join(inp, "ghost__tmp__.xlsx"), HEADERS, [
        ["009", "幽灵", "幽灵部", "火星", 999],
    ])
    res = run_split(_base_cfg(inp, out), log_fn=lambda m: None)
    assert not os.path.exists(os.path.join(res, "幽灵部.xlsx"))   # 残留临时文件被跳过
    assert os.path.exists(os.path.join(res, "销售部.xlsx"))       # 正常文件照常处理


def test_column_mismatch_warns(tmp_path):
    """F7：同一取值跨文件合并、列数不一致时给出预警（不阻断）。"""
    inp = tmp_path / "in"; out = tmp_path / "out"
    inp.mkdir(); out.mkdir()
    _make_book_custom(inp / "A.xlsx", HEADERS, [["001", "张三", "销售部", "北京", 100]])
    _make_book_custom(inp / "B.xlsx", HEADERS + ["备注"],
                      [["002", "李四", "销售部", "上海", 200, "x"]])      # 多一列
    logs = []
    run_split(_base_cfg(str(inp), str(out)), log_fn=logs.append)
    assert any("列数" in m for m in logs), logs


def test_zip_bundles_summary_with_person(two_books):
    """F5：merge=True + 到人 时，每个分组的 ZIP 内应同时含『汇总』与『到人』。"""
    import zipfile
    inp, out = two_books
    res = run_split(_base_cfg(inp, out, to_person=True, person_column="姓名"),
                    log_fn=lambda m: None)   # merge 默认 True → 汇总走扁平文件
    zpath = os.path.join(res, "销售部.zip")
    assert os.path.exists(zpath)
    names = [n.replace("\\", "/") for n in zipfile.ZipFile(zpath).namelist()]
    assert any(n.startswith("汇总/") and n.endswith("销售部.xlsx") for n in names), names   # 汇总进包
    assert any(n.startswith("到人/") and n.endswith("张三.xlsx") for n in names), names      # 到人也在包里


def test_dedupe_path_helper():
    """F1 工具：撞名时追加 (2)/(3)... 后缀。"""
    base = os.path.join("out", "a.xlsx")
    used = set()
    p1 = _dedupe_path(base, used)
    used.add(os.path.normcase(os.path.abspath(p1)))
    p2 = _dedupe_path(base, used)
    assert os.path.basename(p1) == "a.xlsx"
    assert os.path.basename(p2) == "a(2).xlsx"


def test_safe_sheet_title_helper():
    """F3 工具：非法字符替换、超长截断、重名去重。"""
    used = set()
    assert _safe_sheet_title("数据:报表*1", used) == "数据_报表_1"
    assert len(_safe_sheet_title("标" * 40, used)) <= 31
    used2 = set()
    assert _safe_sheet_title("X", used2) == "X"
    assert _safe_sheet_title("X", used2) == "X(2)"
