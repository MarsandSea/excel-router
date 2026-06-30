"""
生成 GIF 演示用样本 — A分公司1-5月员工业绩明细表（脱敏虚拟数据）

表头采用「3行复杂样式」：
  行1  大标题（全列合并，深蓝底色，白粗体 16pt）
  行2  分组标签（基本信息/业绩数据/备注，中蓝底）
  行3  列名（浅蓝底，粗体）
数据从第4行起，末行为合计（淡黄底）。

字段：工号|姓名|所属部门|岗位|区域|销售额(元)|完成率(%)|奖金系数|备注

Copyright (c) 2026 Abelin
MIT License
"""

import os
import random

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

# ── 虚拟员工库（55人） ──────────────────────────────────────────────────────
# 格式：(工号, 姓名, 所属部门, 岗位, 区域, 月度销售目标)
EMPLOYEES = [
    # 销售一部  15人
    ("EMP001", "张伟", "销售一部", "销售经理", "华北", 180000),
    ("EMP002", "李娜", "销售一部", "高级销售", "华北", 150000),
    ("EMP003", "王芳", "销售一部", "销售专员", "华北", 120000),
    ("EMP004", "刘洋", "销售一部", "销售专员", "华北", 120000),
    ("EMP005", "陈静", "销售一部", "销售专员", "东北", 100000),
    ("EMP006", "杨帆", "销售一部", "销售专员", "东北", 100000),
    ("EMP007", "赵磊", "销售一部", "初级销售", "华北",  80000),
    ("EMP008", "周鑫", "销售一部", "初级销售", "东北",  80000),
    ("EMP009", "吴梅", "销售一部", "初级销售", "华北",  80000),
    ("EMP010", "徐刚", "销售一部", "销售专员", "东北", 110000),
    ("EMP011", "孙燕", "销售一部", "销售专员", "华北", 110000),
    ("EMP012", "马超", "销售一部", "高级销售", "东北", 140000),
    ("EMP013", "胡萍", "销售一部", "销售专员", "华北", 105000),
    ("EMP014", "郭亮", "销售一部", "初级销售", "东北",  75000),
    ("EMP015", "林峰", "销售一部", "销售专员", "华北", 115000),
    # 销售二部  15人
    ("EMP016", "黄涛", "销售二部", "销售经理", "华南", 180000),
    ("EMP017", "罗丽", "销售二部", "高级销售", "华南", 150000),
    ("EMP018", "高博", "销售二部", "销售专员", "华东", 120000),
    ("EMP019", "梁红", "销售二部", "销售专员", "华南", 120000),
    ("EMP020", "谢强", "销售二部", "销售专员", "华东", 100000),
    ("EMP021", "宋娟", "销售二部", "销售专员", "华南", 100000),
    ("EMP022", "唐明", "销售二部", "初级销售", "华东",  80000),
    ("EMP023", "韩雪", "销售二部", "初级销售", "华南",  80000),
    ("EMP024", "冯健", "销售二部", "初级销售", "华东",  80000),
    ("EMP025", "秦天", "销售二部", "销售专员", "华南", 110000),
    ("EMP026", "苏丽", "销售二部", "销售专员", "华东", 110000),
    ("EMP027", "蒋军", "销售二部", "高级销售", "华南", 140000),
    ("EMP028", "邓萍", "销售二部", "销售专员", "华东", 105000),
    ("EMP029", "于波", "销售二部", "初级销售", "华南",  75000),
    ("EMP030", "程勇", "销售二部", "销售专员", "华东", 115000),
    # 市场部    10人（非直销岗位，销售额为 0）
    ("EMP031", "曹欣", "市场部", "市场经理", "华北", 0),
    ("EMP032", "彭飞", "市场部", "市场专员", "华南", 0),
    ("EMP033", "朱燕", "市场部", "市场专员", "华东", 0),
    ("EMP034", "丁明", "市场部", "市场专员", "华北", 0),
    ("EMP035", "傅雷", "市场部", "品牌专员", "华南", 0),
    ("EMP036", "沈洁", "市场部", "品牌专员", "华东", 0),
    ("EMP037", "吕强", "市场部", "市场专员", "华北", 0),
    ("EMP038", "石英", "市场部", "市场助理", "华南", 0),
    ("EMP039", "熊伟", "市场部", "市场助理", "华东", 0),
    ("EMP040", "薛涛", "市场部", "市场专员", "华北", 0),
    # 技术支持部  8人
    ("EMP041", "侯斌", "技术支持部", "技术经理", "华北", 0),
    ("EMP042", "段浩", "技术支持部", "技术专员", "华南", 0),
    ("EMP043", "廖玲", "技术支持部", "技术专员", "华东", 0),
    ("EMP044", "范志", "技术支持部", "技术专员", "华北", 0),
    ("EMP045", "钱伟", "技术支持部", "技术专员", "华南", 0),
    ("EMP046", "易鹏", "技术支持部", "技术助理", "华东", 0),
    ("EMP047", "方磊", "技术支持部", "技术助理", "华北", 0),
    ("EMP048", "龙梅", "技术支持部", "技术助理", "华南", 0),
    # 客服部    7人
    ("EMP049", "翁敏", "客服部", "客服经理", "华北", 0),
    ("EMP050", "蔡丽", "客服部", "客服专员", "华南", 0),
    ("EMP051", "毛刚", "客服部", "客服专员", "华东", 0),
    ("EMP052", "任磊", "客服部", "客服专员", "华北", 0),
    ("EMP053", "骆英", "客服部", "客服专员", "华南", 0),
    ("EMP054", "杜波", "客服部", "客服助理", "华东", 0),
    ("EMP055", "纪明", "客服部", "客服助理", "华北", 0),
]

# ── 颜色主题 ─────────────────────────────────────────────────────────────────
C_TITLE    = "1F3864"   # 深藏蓝，行1
C_GROUP    = "2E75B6"   # 中蓝，   行2
C_HEADER   = "BDD7EE"   # 浅蓝，   行3
C_SUBTOTAL = "FFF2CC"   # 淡黄，   合计行
C_WHITE    = "FFFFFF"

COLS = ["工号", "姓名", "所属部门", "岗位", "区域", "销售额(元)", "完成率(%)", "奖金系数", "备注"]
N    = len(COLS)   # 9列
COL_WIDTHS = [10, 8, 14, 12, 8, 14, 12, 10, 18]


def _thin():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _gen_rows(month_num, seed):
    """按月份+seed 生成可复现的数据行（含末行合计）。"""
    rng = random.Random(seed)
    rows = []
    total = 0
    for eid, name, dept, post, area, target in EMPLOYEES:
        if target == 0:
            rows.append([eid, name, dept, post, area, 0, "", "", ""])
        else:
            rate_pct = rng.randint(65, 132)
            sales    = round(target * rate_pct / 100 / 100) * 100
            bonus    = round(max(0.5, min(2.0, rate_pct / 100)), 1)
            if rate_pct >= 120:
                note = "超额完成"
            elif rate_pct >= 100:
                note = ""
            elif rate_pct >= 80:
                note = "基本完成"
            else:
                note = "未达标"
            total += sales
            rows.append([eid, name, dept, post, area, sales, f"{rate_pct}%", bonus, note])
    rows.append(["", "合计", "", "", "", total, "", "", ""])
    return rows


def _make(path, month_num, year=2026):
    wb = Workbook()
    ws = wb.active
    ws.title = "销售明细"

    last_col = get_column_letter(N)

    # ── 行1：大标题（全列合并） ───────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value     = f"北京A分公司  {year}年{month_num}月  员工销售业绩明细表"
    c.font      = Font(size=16, bold=True, color=C_WHITE)
    c.fill      = _fill(C_TITLE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # ── 行2：分组标签 ─────────────────────────────────────────────────────────
    ws.merge_cells("A2:E2")
    ws["A2"].value     = "基本信息"
    ws["A2"].font      = Font(bold=True, color=C_WHITE)
    ws["A2"].fill      = _fill(C_GROUP)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("F2:H2")
    ws["F2"].value     = "业绩数据"
    ws["F2"].font      = Font(bold=True, color=C_WHITE)
    ws["F2"].fill      = _fill(C_GROUP)
    ws["F2"].alignment = Alignment(horizontal="center", vertical="center")

    ws["I2"].value     = "备注"
    ws["I2"].font      = Font(bold=True, color=C_WHITE)
    ws["I2"].fill      = _fill(C_GROUP)
    ws["I2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # ── 行3：列名 ─────────────────────────────────────────────────────────────
    for ci, col_name in enumerate(COLS, 1):
        c            = ws.cell(row=3, column=ci, value=col_name)
        c.font       = Font(bold=True)
        c.fill       = _fill(C_HEADER)
        c.alignment  = Alignment(horizontal="center", vertical="center")
        c.border     = _thin()
    ws.row_dimensions[3].height = 18

    # ── 数据行（第4行起） ─────────────────────────────────────────────────────
    rows = _gen_rows(month_num, seed=year * 100 + month_num)
    for ri, row in enumerate(rows, start=4):
        is_total = (row[1] == "合计")
        for ci, val in enumerate(row, 1):
            c        = ws.cell(row=ri, column=ci, value=val)
            c.border = _thin()
            if is_total:
                c.fill      = _fill(C_SUBTOTAL)
                c.font      = Font(bold=True)
                c.alignment = Alignment(
                    horizontal="right" if ci == 6 else "center"
                )

    # ── 列宽 + 冻结表头 ───────────────────────────────────────────────────────
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"   # 滚动时前3行（含3层表头）始终可见

    wb.save(path)
    print(f"  [OK] {os.path.basename(path)}"
          f"  ({len(rows)-1} 名员工 + 合计行，共 {len(rows)+3} 行)")


def main():
    print(f"正在生成 {len(EMPLOYEES)} 名虚拟员工 × 5 个月份的明细表 ...\n")
    for m in range(1, 6):
        _make(os.path.join(HERE, f"{m}月A分公司明细.xlsx"), m)
    print(
        "\n全部生成完毕，共 5 个文件，位于 examples/\n\n"
        "演示建议：\n"
        "  1. 输入：选 examples/ 文件夹\n"
        "  2. 主拆分列 → 所属部门    （5 个部门分组 × 5 月合并）\n"
        "  3. 开启到人 → 到人列选「姓名」（55 人各自汇总）\n"
        "  4. 高级设置 → 跨文件合并 开，每组 ZIP 开\n"
        "  → 拆出 5 个部门文件夹，每文件夹含汇总+到人，整包 ZIP 可直接发负责人"
    )


if __name__ == "__main__":
    main()
