"""
Excel 智能拆分工具 - 核心拆分逻辑

把一批 Excel 表格按「某一列的取值」拆成多个文件，并完整保留原始表头格式。
面向「真实世界脏数据」：表头不在第一行、多 sheet、xls、公式列、单元格样式各异等都尽量稳。

输入可为单个文件或整个目录；一次运行产出两套结果：
  · 汇总（始终）：按 split_column 拆。单文件/merge=True → 扁平合并；目录+merge=False → 原文件拆分。
  · 到人（可选）：to_person 开 + 有 person_column + 文件名命中 person_file_filter 时，
    再按 person_column 拆到人（同一人跨文件合并到一个文件）。
拆分列（通用，不绑定任何业务）：
  · split_column     —— 主拆分列，必填，按列名选择
  · person_column    —— 到人列，选填
  · selected_values  —— 选填，只拆这些主取值；留空则自动枚举该列所有取值

表头识别（由 config['header_mode'] 控制）：
  · auto    —— 通用启发式自动找表头行（默认）
  · row     —— 用户指定 header_row（1 基行号）
  · keyword —— 旧关键词法：表头需同时含 grid_keys + id_keys（专家兜底）

数据行处理（由 config['preserve_format'] 控制）：
  · 保留格式（默认）：逐行复制值 + 完整单元格格式（字体/颜色/边框/数字格式）
  · 快速模式：数据行只写值，不处理格式，最快

【跨文件合并】同一个取值若出现在多个源文件里，会**追加合并到同一个输出文件**
（config['merge_across_files']，默认 True），从根本上修掉旧版「同名文件互相覆盖」的 bug。
实现：输出工作簿在内存里按取值累积、最后统一保存；第一个产生该取值的源文件定义表头/格式，
后续文件按列位置追加。

【重要】保留格式时，按【原始行号】从源表读取、加法式写入输出，绝不删行——
从根本上规避历史上「删行后用错位行号索引」的 bug。

Copyright (c) 2026 Abelin
MIT License
"""

import os
import re
import time
import shutil
import tempfile
import warnings
import datetime
from copy import copy

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell

from core.utils import soft_clean, hard_clean, apply_alias, is_skip_value, safe_filename

warnings.filterwarnings('ignore')


# =====================================================
# 表头识别
# =====================================================

def _looks_numeric(v):
    """单元格内容看起来像数字（含带千分位/百分号/货币符号的文本）。"""
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    s = soft_clean(v)
    if s == "":
        return False
    s = s.replace(',', '').replace('%', '').replace('¥', '').replace('￥', '').replace('$', '')
    try:
        float(s)
        return True
    except ValueError:
        return False


def detect_header_row_auto(rows, max_scan=15):
    """通用启发式：从前若干行的值矩阵里挑「最像表头的行」。

    表头行的特征：非空单元格多、内容多为文字标签（而非数字）、且下一行起开始出现数据。
    rows 是「行 -> 该行各列值」的二维列表。返回 1 基行号，找不到返回 -1。
    """
    best_idx, best_score = -1, -1.0
    limit = min(len(rows), max_scan)
    for r in range(limit):
        cells = rows[r]
        nonempty = [c for c in cells if soft_clean(c) != ""]
        if not nonempty:
            continue
        n = len(nonempty)
        # 文字占比：表头通常是文字标签，少有纯数字
        str_like = sum(1 for c in nonempty if not _looks_numeric(c))
        str_ratio = str_like / n
        # 数据感：下一行非空单元里数字占比越高，越说明这行是表头
        data_hint = 0.0
        if r + 1 < len(rows):
            nxt = [c for c in rows[r + 1] if soft_clean(c) != ""]
            if nxt:
                data_hint = sum(1 for c in nxt if _looks_numeric(c)) / len(nxt)
        score = n * (0.5 + 0.5 * str_ratio) + n * 0.3 * data_hint
        if score > best_score:
            best_score, best_idx = score, r
    return best_idx + 1 if best_idx >= 0 else -1


def find_header_row(rows, grid_keys, id_keys, max_scan=15):
    """关键词法：在前若干行中找同时包含「网格关键词」和「工号关键词」的表头行。

    rows 为值矩阵，返回 1 基行号，找不到返回 -1。grid_keys/id_keys 任一为空则不启用该条件。
    """
    for r in range(min(len(rows), max_scan)):
        row_str = "".join(hard_clean(c) for c in rows[r])
        ok_grid = (not grid_keys) or any(hard_clean(k) in row_str for k in grid_keys)
        ok_id = (not id_keys) or any(hard_clean(k) in row_str for k in id_keys)
        if ok_grid and ok_id:
            return r + 1
    return -1


def resolve_header_row(rows, config):
    """按 config['header_mode'] 分发：auto / row(指定行号) / keyword。返回 1 基行号或 -1。"""
    mode = config.get('header_mode', 'auto')
    if mode == 'row':
        try:
            hr = int(config.get('header_row', 1))
        except (TypeError, ValueError):
            return -1
        return hr if hr >= 1 else -1
    if mode == 'keyword':
        return find_header_row(rows, config.get('grid_keys', []), config.get('id_keys', []))
    return detect_header_row_auto(rows)


# =====================================================
# 未计算公式检测（防止公式列静默变空白）
# =====================================================

def detect_uncalculated_formulas(work_path, max_rows=200):
    """检测文件是否含「未计算的公式」：有公式(=开头)但没有缓存值。

    真实 Excel 文件保存时会缓存公式计算结果，pandas 能读到真实值；若文件由程序生成或
    公式从未被 Excel 计算过，缓存值为空，拆分后这些列会变成空白且用户不会察觉。
    只采样前 max_rows 行，避免对大文件做双倍全量加载。返回 True 表示存在风险。
    """
    try:
        wb_f = openpyxl.load_workbook(work_path, read_only=True, data_only=False)
        wb_v = openpyxl.load_workbook(work_path, read_only=True, data_only=True)
        risky = False
        for sn in wb_f.sheetnames:
            if sn not in wb_v.sheetnames:
                continue
            ws_f, ws_v = wb_f[sn], wb_v[sn]
            count = 0
            for row_f, row_v in zip(ws_f.iter_rows(), ws_v.iter_rows()):
                for cf, cv in zip(row_f, row_v):
                    if isinstance(cf.value, str) and cf.value.startswith('=') and cv.value is None:
                        risky = True
                        break
                if risky:
                    break
                count += 1
                if count >= max_rows:
                    break
            if risky:
                break
        wb_f.close()
        wb_v.close()
        return risky
    except Exception:
        return False


# =====================================================
# xls 转 xlsx
# =====================================================

def normalize_to_xlsx(file_path, log_fn=None):
    """把 .xls 转成临时 .xlsx 以便 openpyxl 处理；返回 (工作路径, 是否为临时文件)。"""
    if not file_path.lower().endswith('.xls'):
        return file_path, False

    # 临时文件写到系统临时目录（而非源文件同目录），避免残留被下次运行的 os.walk 重复处理。
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        dfs = pd.read_excel(file_path, sheet_name=None, header=None, engine='xlrd')
        with pd.ExcelWriter(tmp_path, engine='openpyxl') as w:
            for sn, df in dfs.items():
                df.to_excel(w, sheet_name=sn, index=False, header=False)
        if log_fn:
            log_fn("  ⚙️ xls 已转换为 xlsx 处理")
        return tmp_path, True
    except Exception as e:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass
        if log_fn:
            log_fn(f"  ⚠️ xls 转换失败：{e}")
        return file_path, False


# =====================================================
# 样式缓存（保留格式模式提速的关键）
# =====================================================

def _cached_copy(style_obj, cache):
    """按源样式对象的 id 缓存复制结果，避免对相同样式重复 copy()。

    真实表格里大量单元格共享同一套样式，openpyxl 内部会复用同一个样式对象（id 相同）。
    缓存后每种样式只 copy 一次，逐行复制也很快。注意：缓存按「单个源文件」作用域使用，
    源文件关闭后弃用，避免不同文件的样式对象 id 复用导致取到旧样式。
    """
    k = id(style_obj)
    cached = cache.get(k)
    if cached is None:
        cached = copy(style_obj)
        cache[k] = cached
    return cached


# =====================================================
# 列名 / 取值枚举（供 GUI 下拉与多选）
# =====================================================

def _read_all_sheets(file_path, config, log_fn=None, nrows=None):
    """读入文件所有 sheet 为「值 DataFrame」字典（header=None）。返回 (df_dict, work_path, is_tmp)。"""
    work_path, is_tmp = normalize_to_xlsx(file_path, log_fn)
    engine = 'openpyxl'
    df_dict = pd.read_excel(work_path, sheet_name=None, header=None,
                            engine=engine, nrows=nrows)
    return df_dict, work_path, is_tmp


def _header_values(df, h_idx):
    """取表头行的列名列表（soft_clean 后）。"""
    header = df.iloc[h_idx - 1]
    return [soft_clean(header.iloc[c]) for c in range(len(header))]


def _find_col_index(df, h_idx, column):
    """按列名在表头行里定位 0 基列索引：先精确匹配，再包含匹配。找不到返回 None。"""
    target = soft_clean(column)
    if not target:
        return None
    names = _header_values(df, h_idx)
    for c, name in enumerate(names):
        if name == target:
            return c
    for c, name in enumerate(names):
        if name and (target in name or name in target):
            return c
    return None


def list_columns(file_path, config, log_fn=None):
    """返回首个可识别 sheet 的列名列表（供 GUI 主/二级列下拉）。只读前若干行，快。"""
    df_dict, work_path, is_tmp = _read_all_sheets(file_path, config, log_fn, nrows=20)
    try:
        for sn, df in df_dict.items():
            rows = df.values.tolist()
            h = resolve_header_row(rows, config)
            if h != -1 and h <= len(df):
                return [c for c in _header_values(df, h) if c]
        return []
    finally:
        if is_tmp and os.path.exists(work_path):
            os.remove(work_path)


def list_values(file_path, config, column, log_fn=None):
    """返回某列去重、归并、去跳过值后的取值列表（供 GUI「只拆这些值」多选）。"""
    alias = config.get('value_alias_map', {})
    skip = config.get('skip_values', [])
    df_dict, work_path, is_tmp = _read_all_sheets(file_path, config, log_fn)
    try:
        values = set()
        for sn, df in df_dict.items():
            head_rows = df.values[:20].tolist()
            h = resolve_header_row(head_rows, config)
            if h == -1 or h > len(df):
                continue
            ci = _find_col_index(df, h, column)
            if ci is None:
                continue
            for v in df.iloc[h:, ci]:
                nv = apply_alias(v, alias)
                if not is_skip_value(nv, skip):
                    values.add(nv)
        return sorted(values)
    finally:
        if is_tmp and os.path.exists(work_path):
            os.remove(work_path)


# =====================================================
# 表头格式读取 / 写入
# =====================================================

def _read_header_format(ws, h_idx):
    """从源 sheet 读取表头区域（1..h_idx 行）的值与格式。返回 (cells, merges, widths)。"""
    cells = []
    for r in range(1, h_idx + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            is_m = isinstance(cell, MergedCell)
            cells.append({
                'row': r, 'col': c,
                'value': None if is_m else cell.value,
                'is_merged': is_m,
                'font':          copy(cell.font)      if not is_m and cell.has_style else None,
                'fill':          copy(cell.fill)      if not is_m and cell.has_style else None,
                'border':        copy(cell.border)    if not is_m and cell.has_style else None,
                'alignment':     copy(cell.alignment) if not is_m and cell.has_style else None,
                'number_format': cell.number_format   if not is_m else None,
            })
    merges = [str(mr) for mr in ws.merged_cells.ranges if mr.max_row <= h_idx]
    widths = {col: ws.column_dimensions[col].width for col in ws.column_dimensions}
    return cells, merges, widths


def _write_header(ws_out, header_cells, merges, widths):
    """把表头的值 + 格式 + 合并单元格 + 列宽写入输出 sheet。"""
    for ci in header_cells:
        if ci['is_merged']:
            continue
        dst = ws_out.cell(row=ci['row'], column=ci['col'])
        dst.value = ci['value']
        if ci['font']:          dst.font          = ci['font']
        if ci['fill']:          dst.fill          = ci['fill']
        if ci['border']:        dst.border        = ci['border']
        if ci['alignment']:     dst.alignment     = ci['alignment']
        if ci['number_format']: dst.number_format = ci['number_format']
    for mr_str in merges:
        ws_out.merge_cells(mr_str)
    for col_letter, width in widths.items():
        if width is not None:
            ws_out.column_dimensions[col_letter].width = width


# =====================================================
# 文件名 / sheet 名去重与净化
# =====================================================

_INVALID_SHEET_CHARS = re.compile(r'[\[\]:*?/\\]')


def _safe_sheet_title(name, used_titles):
    """把任意 sheet 名转成合法且在本工作簿内唯一的 Excel sheet 标题。

    Excel 规则：禁止 [ ] : * ? / \\，不能为空，最长 31 字符，且不能重复。
    used_titles 收集已用的小写标题；重名时在末尾追加 (2)/(3)... 并保证不超 31 字符。
    """
    t = _INVALID_SHEET_CHARS.sub('_', soft_clean(name))[:31] or "Sheet"
    base = t
    i = 2
    while t.lower() in used_titles:
        suffix = f"({i})"
        t = base[:31 - len(suffix)] + suffix
        i += 1
    used_titles.add(t.lower())
    return t


def _dedupe_path(save_path, used_paths):
    """若目标路径已被占用，则在扩展名前追加 (2)/(3)... 直到不冲突。

    用于避免「两个不同取值净化后同名 → 同一 save_path → 后者覆盖前者」的数据丢失。
    used_paths 是已占用路径的规范化集合（os.path.normcase + abspath）。
    """
    norm = os.path.normcase(os.path.abspath(save_path))
    if norm not in used_paths:
        return save_path
    root, ext = os.path.splitext(save_path)
    i = 2
    while True:
        cand = f"{root}({i}){ext}"
        if os.path.normcase(os.path.abspath(cand)) not in used_paths:
            return cand
        i += 1


# =====================================================
# 输出工作簿（按取值在内存中累积，最后统一保存）
# =====================================================

class _OutputBook:
    """一个输出文件对应的内存工作簿，支持跨源文件追加。"""

    __slots__ = ('save_path', 'wb', 'sheets', '_used_titles')

    def __init__(self, save_path):
        self.save_path = save_path
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.sheets = {}        # sheet_name -> {'ws':, 'next_row':, 'ncols':}
        self._used_titles = set()

    def get_or_create_sheet(self, sheet_name, h_idx, header_cells, merges, widths):
        """取得（或首次建立并写入表头的）输出 sheet 信息。"""
        info = self.sheets.get(sheet_name)
        if info is None:
            ws = self.wb.create_sheet(title=_safe_sheet_title(sheet_name, self._used_titles))
            _write_header(ws, header_cells, merges, widths)
            info = {'ws': ws, 'next_row': h_idx + 1}
            self.sheets[sheet_name] = info
        return info

    def save(self):
        """落盘。无任何 sheet 时不保存，返回 False。"""
        if not self.sheets:
            self.wb.close()
            return False
        os.makedirs(os.path.dirname(self.save_path), exist_ok=True)
        self.wb.save(self.save_path)
        self.wb.close()
        return True


def _append_rows(sheet_info, rows_df, preserve, ws_src, src_max_col, style_cache,
                 heartbeat=None):
    """把过滤后的数据行追加到输出 sheet 的当前游标下。

    rows_df 的 index 是源表的【原始 0 基行号】，excel_row = index + 1。
    preserve=True 时从 ws_src 按原始行号逐格复制「值 + 格式」；否则只写值（来自 pandas）。
    heartbeat(n) 每写入约 200 行回调一次，用于向界面报进度并让出 GIL（保持界面响应）。
    """
    ws_out = sheet_info['ws']
    start = sheet_info['next_row']

    if preserve and ws_src is not None:
        for offset, src_idx in enumerate(rows_df.index):
            excel_row = int(src_idx) + 1
            out_row = start + offset
            for c in range(1, src_max_col + 1):
                src_cell = ws_src.cell(row=excel_row, column=c)
                dst = ws_out.cell(row=out_row, column=c)
                dst.value = src_cell.value
                if not isinstance(src_cell, MergedCell):
                    try:
                        if src_cell.has_style:
                            dst.font          = _cached_copy(src_cell.font, style_cache['f'])
                            dst.fill          = _cached_copy(src_cell.fill, style_cache['fl'])
                            dst.border        = _cached_copy(src_cell.border, style_cache['b'])
                            dst.alignment     = _cached_copy(src_cell.alignment, style_cache['a'])
                            dst.number_format = src_cell.number_format
                    except Exception:
                        pass
            if heartbeat and (offset + 1) % 200 == 0:
                heartbeat(200)
    else:
        for offset, (_, row_series) in enumerate(rows_df.iterrows()):
            out_row = start + offset
            for c_idx, val in enumerate(row_series, start=1):
                if pd.isna(val):
                    continue
                ws_out.cell(row=out_row, column=c_idx).value = val
            if heartbeat and (offset + 1) % 200 == 0:
                heartbeat(200)
    if heartbeat and len(rows_df) % 200:
        heartbeat(len(rows_df) % 200)

    sheet_info['next_row'] = start + len(rows_df)


def _matches_filter(name, keywords):
    """文件名是否命中任一关键词；keywords 为空表示全部命中。"""
    if not keywords:
        return True
    n = soft_clean(name)
    return any(soft_clean(k) and soft_clean(k) in n for k in keywords)


def _summary_key_path(output_root, primary, src_stem, single_file, merge):
    """『汇总』输出键与路径。

    单文件 / merge=True → 扁平 {主取值}.xlsx（跨文件合并到一个）。
    文件夹 + merge=False → {主取值}/汇总/{原文件名}.xlsx（原文件拆分，不合并）。
    """
    p = safe_filename(primary)
    if single_file or merge:
        return ('汇总', primary), os.path.join(output_root, f"{p}.xlsx")
    return ('汇总', primary, src_stem), os.path.join(output_root, p, "汇总", f"{src_stem}.xlsx")


def _person_key_path(output_root, primary, person):
    """『到人』输出键与路径：{主取值}/到人/{姓名}.xlsx（同一人跨文件合并到一个）。"""
    p = safe_filename(primary)
    s = safe_filename(person)
    return ('到人', primary, person), os.path.join(output_root, p, "到人", f"{s}.xlsx")


# =====================================================
# 单文件处理
# =====================================================

def process_file(file_path, rel_path, output_root, config, outputs,
                 single_file=False, log_fn=None, stop_flag=None, tick_fn=None):
    """处理单个文件：始终产出『汇总』；满足条件时附加产出『到人』。返回写入总行数。

    · 汇总：按 split_column 拆。单文件/merge=True → 扁平合并；文件夹+merge=False → 原文件拆分。
    · 到人：仅当 to_person 开、有 person_column、且文件名命中 person_file_filter 时，
      按 split_column → person_column 拆，同一人跨文件合并到一个文件。
    · tick_fn(0..1) 报告本文件内部进度（读取→拆分各阶段），供界面进度条使用。
    """
    preserve_format = config.get('preserve_format', True)
    split_column    = soft_clean(config.get('split_column', ''))
    person_column   = soft_clean(config.get('person_column', ''))
    exact           = config.get('exact_match', True)
    alias           = config.get('value_alias_map', {})
    skip            = config.get('skip_values', [])
    merge           = config.get('merge_across_files', True)
    selected = [v for v in (apply_alias(x, alias) for x in config.get('selected_values', [])) if v]
    selected_set = set(selected)

    if not split_column:
        if log_fn:
            log_fn("  ⚠️ 未指定拆分字段，跳过")
        return 0

    base_name = os.path.basename(file_path)
    do_person = (config.get('to_person', False) and not single_file and bool(person_column)
                 and _matches_filter(base_name, config.get('person_file_filter', [])))

    work_path, is_tmp = normalize_to_xlsx(file_path, log_fn)
    out_base = (base_name[:-4] + ".xlsx") if is_tmp else base_name
    src_stem = safe_filename(os.path.splitext(out_base)[0], "source")

    wb_src = None

    def tick(frac):
        """报告本文件内部进度（0~1），供界面进度条平滑推进。"""
        if tick_fn:
            tick_fn(min(max(frac, 0.0), 1.0))

    try:
        tick(0.02)
        if log_fn:
            log_fn("  ⏳ 正在读取数据…（行数多的大文件这一步最慢，请耐心等待）")
        if not is_tmp and detect_uncalculated_formulas(work_path):
            if log_fn:
                log_fn("  ⚠️ 注意：该文件含未计算的公式，相关列可能为空白，"
                       "建议用 Excel 打开另存后再处理")
        tick(0.10)
        preserve_ok = preserve_format and not is_tmp
        if preserve_format and is_tmp and log_fn:
            log_fn("  ⚠️ .xls 转换后无法保留原始格式（仅保留数据，表头格式也会丢失）")

        # ---------- 读数据 ----------
        try:
            df_dict = pd.read_excel(work_path, sheet_name=None, header=None, engine='openpyxl')
        except Exception as e:
            if log_fn:
                log_fn(f"  ❌ 读取失败：{e}")
            return 0
        tick(0.30)

        # ---------- 逐 sheet 识别表头 ----------
        sheet_headers = {}
        for sn, df in df_dict.items():
            h = resolve_header_row(df.values[:20].tolist(), config)
            if h != -1 and h <= len(df):
                sheet_headers[sn] = h
        if not sheet_headers:
            if log_fn:
                log_fn("  ⚠️ 未找到有效表头，跳过")
            return 0

        # ---------- 读表头格式（一次性打开源工作簿） ----------
        if log_fn:
            log_fn("  ⏳ 正在读取格式…")
        header_meta = {}
        try:
            wb_src = openpyxl.load_workbook(work_path, read_only=False, data_only=True)
            for sn, h in sheet_headers.items():
                if sn in wb_src.sheetnames:
                    header_meta[sn] = _read_header_format(wb_src[sn], h)
        except Exception as e:
            if log_fn:
                log_fn(f"  ❌ 读取格式失败：{e}")
            return 0
        est_rows = sum(max(0, len(df_dict[sn]) - h) for sn, h in sheet_headers.items())
        if log_fn:
            log_fn(f"  ⏳ 读取完成（{len(sheet_headers)} 个 sheet，约 {est_rows} 行数据），开始拆分…")
        tick(0.55)

        # 每个源文件一套样式缓存（源工作簿存活期间 id 稳定）
        style_cache = {'f': {}, 'fl': {}, 'b': {}, 'a': {}}
        total_rows = 0
        touched_keys = set()
        rows_written = 0
        last_hb_log = 0

        def heartbeat(n):
            """每批行让出一次 GIL 给界面线程，并周期性在日志里报心跳。"""
            nonlocal rows_written, last_hb_log
            rows_written += n
            time.sleep(0.001)
            if log_fn and rows_written - last_hb_log >= 2000:
                last_hb_log = rows_written
                log_fn(f"    …已拆分写入 {rows_written} 行")

        def emit(key, save_path, sheet_name, h_idx, rows_df, ws_src, src_max_col):
            nonlocal total_rows
            if len(rows_df) == 0 or sheet_name not in header_meta:
                return
            ob = outputs.get(key)
            if ob is None:
                # 防撞名覆盖：不同取值净化后同名时，自动追加 (2)/(3)... 后缀（F1）
                used = {os.path.normcase(os.path.abspath(o.save_path)) for o in outputs.values()}
                ob = _OutputBook(_dedupe_path(save_path, used))
                outputs[key] = ob
            cells, merges_, widths = header_meta[sheet_name]
            sinfo = ob.get_or_create_sheet(sheet_name, h_idx, cells, merges_, widths)
            first_cols = sinfo.setdefault('ncols', src_max_col)
            if first_cols != src_max_col and log_fn:
                log_fn(f"  ⚠️ 列数与首个来源不一致（{first_cols}→{src_max_col}），"
                       f"按列位置合并可能错位：sheet「{sheet_name}」")
            _append_rows(sinfo, rows_df, preserve_ok, ws_src, src_max_col, style_cache,
                         heartbeat=heartbeat)
            total_rows += len(rows_df)
            touched_keys.add(key)

        # ---------- 逐 sheet 拆分 ----------
        n_sheets = max(1, len(sheet_headers))
        for si, (sn, h) in enumerate(sheet_headers.items()):
            if stop_flag and stop_flag():
                break
            df = df_dict[sn]
            col_idx = _find_col_index(df, h, split_column)
            if col_idx is None:
                continue
            pcol_idx = _find_col_index(df, h, person_column) if do_person else None

            data = df.iloc[h:]
            if len(data) == 0:
                continue
            prim = data.iloc[:, col_idx].map(lambda v: apply_alias(v, alias))
            persons = (data.iloc[:, pcol_idx].map(lambda v: apply_alias(v, alias))
                       if (do_person and pcol_idx is not None) else None)

            targets = selected if selected_set else sorted({v for v in prim if not is_skip_value(v, skip)})
            ws_src = wb_src[sn] if (preserve_ok and sn in wb_src.sheetnames) else None
            src_max_col = ws_src.max_column if ws_src is not None else df.shape[1]

            for ti, pval in enumerate(targets):
                if stop_flag and stop_flag():
                    break
                # 进度按「分组」推进：拆分阶段占本文件的 0.55~0.99
                tick(0.55 + 0.44 * ((si + ti / max(1, len(targets))) / n_sheets))
                if is_skip_value(pval, skip):
                    continue
                pmask = prim.map(lambda x: pval in x) if (selected_set and not exact) else (prim == pval)
                if not pmask.any():
                    continue
                sub = data[pmask]

                # 汇总（始终产出）
                k, p = _summary_key_path(output_root, pval, src_stem, single_file, merge)
                emit(k, p, sn, h, sub, ws_src, src_max_col)

                # 到人（可选附加产出）
                if persons is not None:
                    sub_persons = persons[pmask]
                    for person in sorted({v for v in sub_persons if not is_skip_value(v, skip)}):
                        rows_df = sub[sub_persons == person]
                        if len(rows_df):
                            k2, p2 = _person_key_path(output_root, pval, person)
                            emit(k2, p2, sn, h, rows_df, ws_src, src_max_col)

        if log_fn and total_rows:
            log_fn(f"  ✅ 本文件命中 {total_rows} 行 → 分入 {len(touched_keys)} 个输出文件")
        return total_rows
    finally:
        # 无论正常结束还是中途异常，都释放源工作簿并清理临时文件（F2）
        if wb_src is not None:
            try:
                wb_src.close()
            except Exception:
                pass
        if is_tmp and os.path.exists(work_path):
            try:
                os.remove(work_path)
            except OSError:
                pass


# =====================================================
# 主流程
# =====================================================

def run_split(config, log_fn=None, progress_fn=None, stop_flag=None):
    """拆分入口：input_path 可为单文件或目录。一次运行产出『汇总』+ 可选『到人』，
    最后统一保存，并对每个生成了文件夹的主取值打包 ZIP。返回输出目录。

    progress_fn(0~1) 反映整体真实进度：处理各文件占 0~0.75（含文件内部各阶段），
    保存输出占 0.75~0.98，打包收尾到 1.0。"""
    t0 = time.time()
    input_path  = config["input_path"]
    output_root = config["output_path"]
    make_zip    = config.get('make_zip', True)
    single_file = os.path.isfile(input_path)

    timestamp   = datetime.datetime.now().strftime('%m%d%H%M')
    output_path = os.path.join(output_root, f"{timestamp}结果")
    os.makedirs(output_path, exist_ok=True)

    log_file = os.path.join(output_path, "运行日志.txt")

    def _log(msg):
        try:
            with open(log_file, 'a', encoding='utf-8') as lf:
                lf.write(msg + "\n")
        except Exception:
            pass
        if log_fn:
            log_fn(msg)

    # 收集任务（单文件 / 目录递归）
    tasks = []
    if single_file:
        tasks.append((input_path, os.path.basename(input_path)))
    else:
        for root, dirs, files in os.walk(input_path):
            if os.path.abspath(root).startswith(os.path.abspath(output_root)):
                continue
            for f in files:
                if (f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$')
                        and not f.endswith('__tmp__.xlsx')):   # 跳过旧版可能残留的临时文件
                    fp = os.path.join(root, f)
                    tasks.append((fp, os.path.relpath(fp, input_path)))

    do_person = config.get('to_person', False) and not single_file and bool(
        soft_clean(config.get('person_column', '')))

    total = len(tasks)
    _log(f"共找到 {total} 个文件，开始处理...")
    if total == 0:
        _log("⚠️ 输入里没有找到任何 Excel 文件（.xlsx / .xls），请检查输入路径")
    _log(f"拆分字段：{config.get('split_column', '')}"
         + (f"｜到人：{config.get('person_column', '')}" if do_person else "")
         + f"｜跨文件合并：{'是' if config.get('merge_across_files', True) else '否'}\n")

    outputs = {}
    grand_total = 0
    for i, (fp, rp) in enumerate(tasks):
        if stop_flag and stop_flag():
            _log("⛔ 已停止")
            break
        try:
            size_mb = os.path.getsize(fp) / 1048576
            _log(f"[{i + 1}/{total}] {rp}（{size_mb:.1f} MB）")
        except OSError:
            _log(f"[{i + 1}/{total}] {rp}")

        # 整体进度 = (已完成文件数 + 当前文件内部进度) / 总文件数，处理阶段占 0~0.75。
        # 旧版按 (i+1)/total 报进度，单文件场景一开始就 100%，误导用户以为卡死。
        def _tick(frac, _i=i):
            if progress_fn:
                progress_fn((_i + frac) / total * 0.75)

        _tick(0.0)
        try:
            grand_total += process_file(fp, rp, output_path, config, outputs,
                                        single_file, _log, stop_flag, tick_fn=_tick)
        except Exception as e:
            _log(f"  ❌ 处理失败：{e}")
        _tick(1.0)

    # ---------- 统一保存所有输出文件 ----------
    n_out = len(outputs)
    _log(f"\n正在保存 {n_out} 个输出文件...")
    saved = 0
    primaries = set()
    for j, (key, ob) in enumerate(outputs.items()):
        try:
            if ob.save():
                saved += 1
                primaries.add(key[1])   # key = (tree, primary, ...)
        except Exception as e:
            _log(f"  ❌ 保存失败 {ob.save_path}：{e}")
        if progress_fn:
            progress_fn(0.75 + 0.23 * (j + 1) / n_out)
        if (j + 1) % 10 == 0 and (j + 1) < n_out:
            _log(f"  …已保存 {j + 1}/{n_out}")
        time.sleep(0.002)   # 让出 GIL，保存阶段保持界面响应

    # ---------- 对每个生成了文件夹的主取值打包 ZIP（方便分发） ----------
    if make_zip and not single_file:
        for primary in primaries:
            folder = os.path.join(output_path, safe_filename(primary))
            if not os.path.isdir(folder):
                continue
            # 若该取值的汇总是扁平文件（跨文件合并模式），先复制进 {取值}/汇总/ 再打包，
            # 保证每个分组的 ZIP 自带「汇总 + 到人」，方便整包发负责人。
            sb = outputs.get(('汇总', primary))
            if sb is not None and os.path.exists(sb.save_path):
                try:
                    dst_dir = os.path.join(folder, "汇总")
                    os.makedirs(dst_dir, exist_ok=True)
                    shutil.copy2(sb.save_path, os.path.join(dst_dir, os.path.basename(sb.save_path)))
                except Exception:
                    pass
            if any(os.scandir(folder)):
                try:
                    shutil.make_archive(folder, 'zip', folder)
                    _log(f"  📦 {safe_filename(primary)}.zip")
                except Exception as e:
                    _log(f"  ⚠️ 打包失败 {primary}：{e}")

    if progress_fn:
        progress_fn(1.0)
    mm, ss = divmod(int(time.time() - t0 + 0.5), 60)
    _log(f"\n✅ 全部完成！生成 {saved} 个文件，共写入 {grand_total} 行，"
         f"耗时 {f'{mm} 分 {ss} 秒' if mm else f'{ss} 秒'}")
    _log(f"📁 输出目录：{output_path}")

    return output_path
